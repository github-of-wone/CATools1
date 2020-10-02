#%%

# https://github.com/github-of-wone/CATools1
# Twitter @WoneAdvisers

"""
https://services.gst.gov.in/services/api/search/taxpayerDetails
https://services.gst.gov.in/services/api/search/goodservice?gstin=27AAHCA0613R1ZH
https://services.gst.gov.in/services/api/search/taxpayerReturnDetails
"""

# Crashes when it encounters less than 10 rows in the table
# Need to modularize it
# Need to test mypy and black on it
# why am i getting error ar the bottom to use some other chrome?
# 01SEP2020 now i realise that this could have much simply be done using XHR!
# https://stackoverflow.com/questions/52109777/how-to-wait-for-a-button-to-be-clicked-by-user-in-selenium-web-driver-using-pyth

from os import path

from openpyxl import Workbook, load_workbook
from PySimpleGUI import Cancel, FileBrowse, Input, Ok, Text, Window, theme

# from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


# def Convert(a: list) -> Dictionary:
# Source: https://www.geeksforgeeks.org/python-convert-a-list-to-dictionary/
def Convert_list_to_dictionary(a):
    it = iter(a)
    res_dct = dict(zip(it, it))
    return res_dct


theme("reddit")
event, values = Window(
    "Select the Excel File",
    [
        [
            Text(
                "Select the Excel File with only 'Sheet1' where all GSTINs are in the 1st Row from Cell A2"
            )
        ],
        [Input(), FileBrowse()],
        [Ok(), Cancel()],
    ],
).read(close=True)

# conversationPath = "/content/wa.txt" #if doing on google colab
conversationPath = values["Browse"]

a = path.basename(conversationPath)  # filename
b = path.dirname(conversationPath)  # directory
c = path.splitext(a)[0]

# action_on_file = "TEST.xlsx"  # "GST Filing Data.xlsx"
action_on_file = b + r"/" + c + ".xlsx"

# Selenium part begins --------------------------------------------------------
option = webdriver.ChromeOptions()
option.add_argument(" — incognito")

browser = webdriver.Chrome(
    # executable_path="D:\\Cloud\\OneDrive\\DAY LOG\\~~~GDRIVE\\~~~2020\\25 14JUN2020 SUNDAY\\sel\\chromedriver_win32\\chromedriver.exe",
    executable_path=r"D:\convert\chromedriver85.exe",
    chrome_options=option,
)  # Where to get the Chrome Driver!


wb = load_workbook(filename=action_on_file, read_only=False)
ws_index = wb["INDEX"]
gst_numbers_list = []
for row in ws_index.rows:
    # for cell in row:815
    if row[0].value != None:
        gst_numbers_list.append(row[0].value)
gst_numbers_list.pop(0)


# gst_numbers_list[0] #this is for the first allocation, will have to loop this

#%%
for gstin in gst_numbers_list:

    # gstin = "27AAHCA0613R1ZH"  # Keep the above line when looping
    browser.get("https://services.gst.gov.in/services/searchtp")

    # Wait 20 seconds for page to load
    timeout = 20

    try:
        WebDriverWait(browser, timeout).until(
            EC.visibility_of_element_located((By.XPATH, "//p[@class='m-cir reg']"))
        )

    except TimeoutException:
        print("Timed out waiting for page to load")
        browser.quit()

    # find_elements_by_xpath returns an array of selenium objects.
    gstin_num = browser.find_element_by_xpath(
        "//input[@id='for_gstin']"
    )  # note here the change of find_elements_by_xpath to find_element_by_xpath
    # gstin=input("Enter the GST number and press Enter\n")
    gstin_num.send_keys(gstin)

    layout1 = [[Text("Enter CAPTCHA - ALWAYS ON TOP")], [Input()], [Ok()]]
    window = Window(
        "CAPTCHA ENTRY",
        layout1,
        no_titlebar=False,
        auto_size_buttons=False,
        keep_on_top=True,
        grab_anywhere=True,
    ).read(close=True)

    captcha_text = browser.find_element_by_xpath("//input[@id='fo-captcha']")
    # captcha = input("Enter the captcha and press Enter\n")
    captcha = window[1][0]
    captcha_text.send_keys(captcha)

    # window = sg.Window('Running Timer', layout, no_titlebar=True, auto_size_buttons=False, keep_on_top=True, grab_anywhere=True)

    browser.find_element_by_xpath("//button[@id='lotsearch']").click()

    try:
        WebDriverWait(browser, timeout).until(
            EC.visibility_of_element_located(
                (By.XPATH, "//p[contains(text(),'Nature of Business Activities')]")
            )
        )

    except TimeoutException:
        print("Timed out waiting for page to load")
        browser.quit()

    browser.find_element_by_xpath("//button[@id='filingTable']").click()

    p2_list = browser.find_elements_by_xpath('//*[@id="lottable"]/div[2]')[
        0
    ].text.split("\n")
    tempvar = p2_list[-1]
    part2 = Convert_list_to_dictionary(p2_list)
    if part2["Effective Date of Cancellation"] == "Principal Place of Business":
        part2["Effective Date of Cancellation"] = ""
        part2["Principal Place of Business"] = tempvar

    p3_list = browser.find_elements_by_xpath('//*[@id="lottable"]/div[3]')[
        0
    ].text.split("\n")
    p3_list.pop(0)
    part3 = Convert_list_to_dictionary(p3_list)
    # part3=[x for x in p3_list) if p3_list.index(x)]

    p4_list = browser.find_elements_by_xpath(
        '//*[@id="lottable"]/div[4]/div/div/div/table'
    )[0].text.split("\n")
    p4_list.pop(0)
    p4_list.pop(0)
    part4 = {}
    for counter1 in range(len(p4_list)):
        part4[counter1] = p4_list[counter1]

    # part5 = {}
    # part5_parts = {"GSTR3B": {}, "GSTR1": {}, "GSTR9": {}, "GSTR9C": {}}
    # part5_innerdict_keys = ["Financial Year", "Tax Period", "Date of filing", "Status"]
    tables = browser.find_elements_by_xpath('//*[@data-ng-table="listOfStatus"]')
    # part5_parts = ["GSTR3B","GSTR1","GSTR9","GSTR9C"]

    part6_parts = [[], [], [], []]
    # dictionary1=[1,2,3,4]
    for idx, table in enumerate(tables):
        rows = table.find_elements_by_xpath("tbody/tr")
        # print(idx)
        for row in rows:
            items = row.find_elements_by_xpath("td")
            data1 = [val.text for val in items]
            # print(data1)
            # dictionary1[idx] = dict(zip(part5_innerdict_keys, data1))
            part6_parts[idx].append(data1)

    part6_GSTR3B = part6_parts[0]
    part6_GSTR1 = part6_parts[1]
    part6_GSTR9 = part6_parts[2]
    part6_GSTR9C = part6_parts[3]

    # Append the legal_name_of_business and gstin in this INDEX Sheet
    # ws_index.cell(row=2,column=2,value=gstin) #here need to change the column 1 value too to match

    # Create a new sheet by 15 digits of the legal_name_of_business
    ws1 = wb.create_sheet(part2["Legal Name of Business"][0:16])

    # Insert into this new sheet in cells
    ws1["A1"] = gstin

    ws1["A3"] = "Legal Name of Business"
    ws1["A4"] = part2["Legal Name of Business"]

    ws1["B3"] = "Trade Name"
    ws1["B4"] = part2["Trade Name"]

    ws1["C3"] = "Administrative Office"
    ws1["C4"] = part2["Administrative Office"]

    ws1["A6"] = "Other Administrative Office"
    ws1["A7"] = part2["Other Office"]

    ws1["B6"] = "Date of Registration"
    ws1["B7"] = part2["Date of registration"]

    ws1["C6"] = "Constitution of Business"
    ws1["C7"] = part2["Constitution of Business"]

    ws1["A9"] = "Taxpayer Type"
    ws1["A10"] = part2["Taxpayer Type"]

    ws1["B9"] = "GSTIN or UIN Status"
    # ws1["B10"] = part2["GSTIN / UIN Status"]

    ws1["C9"] = "Effective Date of Cancellation"
    # ws1["C10"] = part2["Effective Date of Cancellation"]

    ws1["A12"] = "Principal Place of Business"
    ws1["A13"] = part2["Principal Place of Business"]

    ws1["A15"] = "Nature of Business Activities"
    # ws1["A16"] = part3  ### TODO SEE WHAT IS BEING SHOWN HERE : gives ValueError
    ws1["A16"] = part3["1."]
    try:
        ws1["B16"] = part3["2."]
    except:
        pass
    try:
        ws1["C16"] = part3["3."]
    except:
        pass
    try:
        ws1["D16"] = part3["4."]
    except:
        pass

    # ws1["A18"] = "Return Type"
    ws1["A21"] = "Financial Year"
    ws1["B21"] = "Tax Period"
    ws1["C21"] = "Date of Filing"
    ws1["D21"] = "Status"
    ws1["E21"] = "GST Return"

    x = ["A", "B", "C", "D"]

    try:
        for idx, val in enumerate(part6_GSTR3B):
            for idx2, val2 in enumerate(val):
                ws1[f"{x[idx2]}{idx+23}"] = val2
            ws1["E" + f"{idx+23}"] = "GSTR-3B"
    except:
        pass

    try:
        for idx, val in enumerate(part6_GSTR1):
            for idx2, val2 in enumerate(val):
                ws1[f"{x[idx2]}{idx+33}"] = val2
            ws1["E" + f"{idx+33}"] = "GSTR-1"
    except:
        pass

    try:
        for idx, val in enumerate(part6_GSTR9):
            for idx2, val2 in enumerate(val):
                ws1[f"{x[idx2]}{idx+43}"] = val2
            ws1["E" + f"{idx+43}"] = "GSTR-9"

        for idx, val in enumerate(part6_GSTR9C):
            for idx2, val2 in enumerate(val):
                ws1[f"{x[idx2]}{idx+45}"] = val2
            ws1["E" + f"{idx+45}"] = "GSTR-9C"
    except:
        pass

    # Save the file and close it
    wb.save(filename=action_on_file)

Window.close
browser.quit()


# %%
