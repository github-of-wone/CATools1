""" This file is for logging into MCA Site and downloading the data as required"""

# https://github.com/github-of-wone/CATools1
# Twitter @WoneAdvisers

#%%
import csv
import time

from lxml.html import fromstring
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# Selenium part begins --------------------------------------------------------
option = webdriver.ChromeOptions()
option.add_argument(" — incognito")

browser = webdriver.Chrome(
    executable_path=r"D:\convert\chromedriver85.exe", chrome_options=option,
)

browser.get("http://www.mca.gov.in/mcafoportal/viewCompanyMasterData.do")

# Wait 20 seconds for page to load
timeout = 30

try:
    WebDriverWait(browser, timeout).until(
        EC.visibility_of_element_located(
            (By.XPATH, "//*[@id='companyLLPMasterData_0']")
        )
    )
except TimeoutException:
    print("Timed out waiting for page to load")
    browser.quit()
#%%
# Process on MCA Site begins ----------------------------------------

# TODO first the CIN of the Company needs to be found out. stored in a variable.  There could be many names that come up
# TODO present the CIN of the company
# TODO put the captcha and click on submit

# TODO Gather the master data that is being presented in a variable # Presently let the excel only be used for the purpose of speed.


#%%
# TODO go to Public documents link and click on the details that are made available and download the document_name and date_of_filing
company_name = "RELIANCE INDUSTRIES LIMITED"
browser.get("http://www.mca.gov.in/mcafoportal/viewPublicDocumentsFilter.do")

browser.find_element_by_xpath("//*[@id='companyChk']").click()
browser.find_element_by_xpath("//*[@id='companyName']").send_keys(
    company_name
)  # Assumption here is that the entity search is of a Company
browser.find_element_by_xpath("//*[@id='viewDocuments_0']").click()

browser.find_element_by_xpath(
    "//*[@id='viewDocuments_0']"
).click()  # Here the 1st Company /LLP name that comes up is selected
browser.find_element_by_xpath("//*[@id='results']/tbody/tr[2]/td[2]/a").click()

#%%
form_data = []
document_category = browser.find_element_by_xpath(
    "//*[@id='viewCategoryDetails_categoryName']"
).text.split("\n")
financial_year = browser.find_element_by_xpath(
    "//*[@id='viewCategoryDetails_finacialYear']"
).text.split("\n")


#%%

alist1 = []

for dn_idx, document_name in enumerate(
    document_category[1:]
):  # Certificates is starting from 2
    # For all forms submitted in document_category <say: Certificates>:
    idx = dn_idx + 2

    alist1.append(list())

    browser.find_element_by_xpath(
        f"//*[@id='viewCategoryDetails_categoryName']/option[{idx}]"
    ).click()  # Here a loop will be required, this is for Document Category

    for year_idx, year_num in enumerate(
        financial_year[1:16]
    ):  # 2006 is starting from 2
        # For all for forms submitted in that document_category in a financial_year <say: 2006>:
        idx2 = year_idx + 2

        browser.find_element_by_xpath(
            f"//*[@id='viewCategoryDetails_finacialYear']/option[{idx2}]"
        ).click()  # Here a loop will be required, this is for Year of Filing

        try:
            browser.find_element_by_xpath("//*[@id='viewCategoryDetails_0']").click()
        except:
            print("rror")
            time.sleep(2)
            browser.find_element_by_xpath("//*[@id='msgboxclose']").click()

        if (
            "No documents are available for the selected category"
            not in browser.page_source
        ):
            # TODO code to have it put into a table format
            data1 = fromstring(browser.page_source)
            z = data1.xpath("//table[@id='results']//tr")

            for row in z[1:]:
                print([td.text.strip() for td in row.xpath(".//td")])
                alist = [td.text.strip() for td in row.xpath(".//td")]
                alist1[dn_idx].append(alist)


# for document_categories in data1.xpath('//*[@id="viewCategoryDetails_categoryName"]'):
# click on submit button, get the whole table (which will include the items which are having style="display:none;")
# add it a list of the format - <Name / CIN of the Company> , <document_category>, <financial_year>, <document_name>, <date_of_filing>

# Output this list in a csv


#%%
wb = Workbook()
for item in range(len(alist1)):
    # ws1 = wb.create_sheet(title=document_name.lstrip())
    ws = wb.create_sheet(str(item))
    for row1 in range(1, len(alist1[item]) + 1):
        for col in range(1, 3):
            _ = ws.cell(column=col, row=row1, value=alist1[item][row1 - 1][col - 1])
wb.save(filename="entity pubdocs on site.xlsx")

#%%
document_category = [
    "    -----Select-----",
    "    Certificates",
    "    Change in Directors",
    "    Incorporation Documents",
    "    Charge Documents",
    "    Annual Returns and Balance Sheet eForms",
    "    LLP Forms(Conversion of company to LLP)",
    "    Other eForm Documents",
    "    Other Attachments",
]

financial_year = [
    "    -----Select-----",
    "    2006",
    "    2007",
    "    2008",
    "    2009",
    "    2010",
    "    2011",
    "    2012",
    "    2013",
    "    2014",
    "    2015",
    "    2016",
    "    2017",
    "    2018",
    "    2019",
    "    2020",
    "    2021",
    "    2022",
    "    2023",
    "    2024",
    "    2025",
    "    2026",
    "    2027",
    "    2028",
    "    2029",
    "    2030",
]

#%%
"""

a = browser.find_element_by_xpath("//*[@id='results']")
#%%
a = browser.find_element_by_xpath(
    "/html/body/div[1]/div[6]/div[1]/section/form[2]/table[1]"
)
#%%



#%%
# table =  browser.find_element_by_xpath("//table[@class='result-forms_vpd']")
# table =  browser.find_element_by_xpath("//*[@id='results']")
# table =  browser.find_element_by_xpath("//table[@id='results']")
# table = browser.find_element_by_xpath("//table[@id='results']")


#%%
### Source:https://stackoverflow.com/questions/37090653/iterating-through-table-rows-in-selenium-python
b = table.find_elements_by_xpath(".//tr")

for row in b:
    # print([td.text for td in row.find_elements_by_xpath(".//td[@class='dddefault'][1]"])

    print([td.text for td in row.find_elements_by_xpath(".//td[1]")])
    print([td.text for td in row.find_elements_by_xpath(".//td[2]")])

#%%
table.find_element_by_xpath("//*[@id='results']/tbody/tr[4]/td[1]").text

"""
#%%

### Extra ---------------------------------------------------------------------

document_category = browser.find_element_by_xpath(
    "//*[@id='viewCategoryDetails_categoryName']"
).text.split("\n")
""" The above will give this
['    -----Select-----',
 '    Certificates',
 '    Change in Directors',
 '    Incorporation Documents',
 '    Charge Documents',
 '    Annual Returns and Balance Sheet eForms',
 '    LLP Forms(Conversion of company to LLP)',
 '    Other eForm Documents',
 '    Other Attachments']
"""

financial_year = browser.find_element_by_xpath(
    "//*[@id='viewCategoryDetails_finacialYear']"
).text.split("\n")
""" The above will give this
['    -----Select-----',
 '    2006',
 '    2007',
 '    2008',
 '    2009',
 '    2010',
 '    2011',
 '    2012',
 '    2013',
 '    2014',
 '    2015',
 '    2016',
 '    2017',
 '    2018',
 '    2019',
 '    2020',
 '    2021',
 '    2022',
 '    2023',
 '    2024',
 '    2025',
 '    2026',
 '    2027',
 '    2028',
 '    2029',
 '    2030']
 """
